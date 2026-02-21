import os
from datetime import datetime

from django.core.management.base import BaseCommand
from django.db import connection, transaction
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from loans.models import Customer, Loan


class Command(BaseCommand):
    help = "Load customer and loan data from Excel files"

    def handle(self, *args, **options):
        base_path = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        )
        customer_file = os.path.join(base_path, "customer_data.xlsx")
        loan_file = os.path.join(base_path, "loan_data.xlsx")

        # Load customers first
        self.stdout.write("Loading customer data...")
        self.load_customers(customer_file)

        # Load loans
        self.stdout.write("Loading loan data...")
        self.load_loans(loan_file)

        # Reset PostgreSQL sequences so new records don't clash with loaded IDs
        self.reset_sequences()

        self.stdout.write(self.style.SUCCESS("Data loading completed successfully!"))

    def reset_sequences(self) -> None:
        """Reset PostgreSQL auto-increment sequences after bulk insert with explicit IDs"""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT setval(
                    pg_get_serial_sequence('loans_customer', 'customer_id'),
                    COALESCE((SELECT MAX(customer_id) FROM loans_customer), 1)
                );
            """
            )
            cursor.execute(
                """
                SELECT setval(
                    pg_get_serial_sequence('loans_loan', 'loan_id'),
                    COALESCE((SELECT MAX(loan_id) FROM loans_loan), 1)
                );
            """
            )
        self.stdout.write(self.style.SUCCESS("Sequences reset successfully!"))

    @transaction.atomic
    def load_customers(self, file_path: str) -> None:
        """Load customers from Excel file"""
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        wb = load_workbook(file_path)
        ws: Worksheet | None = wb.active

        if ws is None:
            self.stdout.write(self.style.ERROR(f"No active worksheet found"))
            return

        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue

            if row[0] is None:  # Skip empty rows
                continue

            try:
                # Excel columns: Customer ID | First Name | Last Name | Age | Phone Number | Monthly Salary | Approved Limit
                customer_id = int(row[0] or 0)  # type: ignore
                first_name = str(row[1] or "").strip()  # type: ignore
                last_name = str(row[2] or "").strip()  # type: ignore
                age = int(row[3] or 25)  # type: ignore
                phone_number = str(row[4] or "").strip()  # type: ignore
                monthly_salary = float(row[5] or 0.0)  # type: ignore
                approved_limit = float(row[6] or 0.0)  # type: ignore
                # current_debt column may or may not exist
                current_debt = float(row[7] or 0.0) if len(row) > 7 and row[7] is not None else 0.0  # type: ignore

                # Skip if customer or phone already exists
                if Customer.objects.filter(customer_id=customer_id).exists():
                    continue
                if Customer.objects.filter(phone_number=phone_number).exists():
                    continue

                Customer.objects.create(
                    customer_id=customer_id,
                    first_name=first_name,
                    last_name=last_name,
                    age=age,
                    phone_number=phone_number,
                    monthly_income=monthly_salary,
                    approved_limit=approved_limit,
                    current_debt=current_debt,
                )
                self.stdout.write(
                    f"Created customer: {customer_id} - {first_name} {last_name}"
                )

            except (ValueError, IndexError, TypeError) as e:
                self.stdout.write(
                    self.style.WARNING(f"Error in row {row_idx}: {str(e)}")
                )
                continue

        wb.close()
        self.stdout.write(self.style.SUCCESS(f"Loaded customers from {file_path}"))

    @transaction.atomic
    def load_loans(self, file_path: str) -> None:
        """Load loans from Excel file"""
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        wb = load_workbook(file_path)
        ws: Worksheet | None = wb.active

        if ws is None:
            self.stdout.write(self.style.ERROR(f"No active worksheet found"))
            return

        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue

            if row[0] is None:  # Skip empty rows
                continue

            try:
                # Safe type conversion for openpyxl cells
                customer_id = int(row[0] or 0)  # type: ignore
                loan_id = int(row[1] or 0)  # type: ignore
                loan_amount = float(row[2] or 0.0)  # type: ignore
                tenure = int(row[3] or 0)  # type: ignore
                interest_rate = float(row[4] or 0.0)  # type: ignore
                monthly_repayment = float(row[5] or 0.0)  # type: ignore
                emis_paid = int(row[6] or 0)  # type: ignore
                start_date = row[7]
                end_date = row[8]

                # Convert string dates to date objects
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

                # Get customer
                try:
                    customer = Customer.objects.get(customer_id=customer_id)
                except Customer.DoesNotExist:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Customer {customer_id} not found in row {row_idx}"
                        )
                    )
                    continue

                # Skip if loan already exists
                if Loan.objects.filter(loan_id=loan_id).exists():
                    continue

                Loan.objects.create(
                    loan_id=loan_id,
                    customer=customer,
                    loan_amount=loan_amount,
                    tenure=tenure,
                    interest_rate=interest_rate,
                    monthly_installment=monthly_repayment,
                    emis_paid=emis_paid,
                    status="approved",  # Past loans are approved
                    start_date=start_date,
                    end_date=end_date,
                )
                self.stdout.write(f"Created loan: {loan_id} - Customer {customer_id}")

            except (ValueError, IndexError, TypeError) as e:
                self.stdout.write(
                    self.style.WARNING(f"Error in row {row_idx}: {str(e)}")
                )
                continue

        wb.close()
        self.stdout.write(self.style.SUCCESS(f"Loaded loans from {file_path}"))
